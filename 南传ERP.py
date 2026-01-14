import requests,re,calendar
import hashlib
import uuid
import json
import time
import ddddocr
from lxml import etree
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json


ocr = ddddocr.DdddOcr()
def create_sales_report_excel(data_list, start_date="2024-1-1", end_date="2024-2-1",
                              filename="销售在途汇总表.xlsx"):
    """
    创建销售汇总Excel报表

    Args:
        data_list: 数据列表
        start_date: 开始日期
        end_date: 结束日期
        filename: 输出文件名
    """
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "销售在途汇总表"

    # 设置样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    subtitle_font = Font(name='微软雅黑', size=12, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='微软雅黑', size=10)

    # 居中对齐
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头填充色
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    # 第一行：主标题
    ws.merge_cells('A1:L1')
    ws['A1'] = f"{start_date}至{end_date} 销售在途汇总表"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    # 第二行：副标题
    current_date = datetime.now().strftime('%Y-%m-%d')
    ws.merge_cells('A2:L2')
    ws['A2'] = f"采集时间 {current_date} 销售在途汇总表"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = center_alignment

    # 第三行：表头
    headers = [
        '业务机构', '客户单位', '品种数', '数量', '码洋(元)',
        '含税金额/实洋(元)', '折扣额(元)', '税额(元)',
        '不含税金额(元)', '未开票数量', '未开票金额', '未开票成本'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # 填充数据
    for row_idx, data in enumerate(data_list, 4):
        for col_idx, (key, value) in enumerate(data.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # 设置对齐方式
            if key in ['业务机构', '客户单位']:
                cell.alignment = left_alignment
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # 特殊格式化：合计行
            if data.get('业务机构') == '合计':
                cell.font = Font(name='微软雅黑', size=10, bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # 设置列宽
    column_widths = [25, 20, 8, 10, 12, 15, 12, 12, 15, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 20

    # 冻结窗格（冻结标题行）
    ws.freeze_panes = 'A4'

    # 保存文件
    wb.save(filename)
    print(f"Excel文件已生成: {filename}")

    return filename


def get_month_first_last(date=None):
    """获取指定日期所在月份的第一天和最后一天"""
    if date is None:
        date = datetime.date.today()

    # 获取月份第一天
    first_day = date.replace(day=1)

    # 获取月份最后一天
    _, last_day_num = calendar.monthrange(date.year, date.month)
    last_day = date.replace(day=last_day_num)

    return first_day, last_day



# first_day, last_day = get_month_first_last()

def home_req(session):
    """获取初始cookie和XSRF-TOKEN"""
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "Connection": "keep-alive",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "sec-ch-ua": "\"Chromium\";v=\"142\", \"Microsoft Edge\";v=\"142\", \"Not_A Brand\";v=\"99\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\""
    }
    url = "https://uat-publish.nfcb.com.cn/oauth/login"

    try:
        response = session.get(url, headers=headers)
        cookies_dict = response.cookies.get_dict()

        # 调试信息

        if not cookies_dict:
            print("[ERROR] home_req - 未获取到cookie")
            return None

        return cookies_dict

    except Exception as e:
        print(f"[ERROR] home_req - 请求失败: {e}")
        return None


def get_images(cookies_josn, session):
    """获取验证码"""
    headers = {
        "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "Connection": "keep-alive",
        "Referer": "https://uat-publish.nfcb.com.cn/oauth/login",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
    }
    url = f"https://uat-publish.nfcb.com.cn/code/image"

    try:
        response = session.get(url, headers=headers, cookies=cookies_josn)

        # 调试信息


        if response.status_code != 200 or len(response.content) == 0:
            print("[ERROR] get_images - 获取验证码失败")
            return None

        result = ocr.classification(response.content)
        print(f"[DEBUG] get_images - 识别结果: {result}")
        return result

    except Exception as e:
        print(f"[ERROR] get_images - 请求失败: {e}")
        return None


def org_id(username, cookies_josn, session):
    """获取机构ID"""
    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "Connection": "keep-alive",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Origin": "https://uat-publish.nfcb.com.cn",
        "Referer": "https://uat-publish.nfcb.com.cn/oauth/login",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "X-Requested-With": "XMLHttpRequest",
        # 动态获取XSRF-TOKEN
        "X-XSRF-TOKEN": cookies_josn.get("XSRF-TOKEN", ""),
        "sec-ch-ua": "\"Chromium\";v=\"142\", \"Microsoft Edge\";v=\"142\", \"Not_A Brand\";v=\"99\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\""
    }
    url = "https://uat-publish.nfcb.com.cn/organ/queryByUserName"
    data = {
        "username": username
    }

    try:
        response = session.post(url, headers=headers, data=data, cookies=cookies_josn)

        # 调试信息


        if response.status_code != 200:
            print(f"[ERROR] org_id - 请求失败，状态码: {response.status_code}")
            return None

        result = response.json()
        if result.get("data"):
            organ_id = result["data"][0]["id"]
            print(f"[DEBUG] org_id - 获取到机构ID: {organ_id}")
            return organ_id
        else:
            print(f"[ERROR] org_id - 响应格式错误: {result}")
            return None

    except Exception as e:
        print(f"[ERROR] org_id - 请求失败: {e}")
        return None


def login(imageCode, cookies_josn, organ, username, session):
    """执行登录"""
    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "Connection": "keep-alive",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Origin": "https://uat-publish.nfcb.com.cn",
        "Referer": "https://uat-publish.nfcb.com.cn/oauth/login",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "X-Requested-With": "XMLHttpRequest",
        "X-XSRF-TOKEN": cookies_josn.get("XSRF-TOKEN", ""),
        "sec-ch-ua": "\"Chromium\";v=\"142\", \"Microsoft Edge\";v=\"142\", \"Not_A Brand\";v=\"99\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\""
    }

    url = "https://uat-publish.nfcb.com.cn/authentication/form"
    data = {
        "username": username,
        "organ": organ,
        "password": hashlib.sha256('Erp123456'.encode('utf-8')).hexdigest(),
        "imageCode": imageCode,
        "passwordLevel": "true"
    }

    try:

        response = session.post(url, headers=headers, data=data, cookies=cookies_josn)

        # 调试信息


        if response.status_code == 200:
            result = response.json()
            if result.get("msg")=="登录成功":
                print(f"[SUCCESS] 用户 {username} 登录成功")
                return True
            else:
                print(f"[ERROR] 用户 {username} 登录失败: {result.get('message', '未知错误')}")
                return False
        else:
            print(f"[ERROR] 登录请求失败，状态码: {response.status_code}")
            return False

    except Exception as e:
        print(f"[ERROR] login - 请求失败: {e}")
        return False


def get_token_index(session):
    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "Connection": "keep-alive",
        "Referer": "https://uat-publish.nfcb.com.cn/manager/center",
        "Sec-Fetch-Dest": "iframe",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "sec-ch-ua": "\"Chromium\";v=\"142\", \"Microsoft Edge\";v=\"142\", \"Not_A Brand\";v=\"99\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\""
    }

    url = "https://uat-publish.nfcb.com.cn/manager/report/sale/rtSaleUnderwaySumReport/index"
    response = session.get(url, headers=headers)

    pattern1 = r"const _token = '([^']+)'"
    match1 = re.search(pattern1, response.text)
    print(match1.group(1))

    return match1.group(1)



def process_user(username):
    """处理单个用户的登录流程"""
    print(f"\n{'=' * 50}")
    print(f"开始处理用户: {username}")
    print(f"{'=' * 50}")

    # 为每个用户创建独立的Session
    session = requests.Session()

    try:
        # 1. 获取初始cookie
        cookies_josn = home_req(session)

        if not cookies_josn:
            print(f"[ERROR] 用户 {username} 无法获取初始cookie，跳过")
            return False

        # 2. 获取验证码
        imageCode = get_images(cookies_josn, session)

        if not imageCode:
            print(f"[ERROR] 用户 {username} 无法获取验证码，跳过")
            return False

        # 3. 获取机构ID
        organ = org_id(username, cookies_josn, session)

        if not organ:
            print(f"[ERROR] 用户 {username} 无法获取机构ID，跳过")
            return False

        # 4. 执行登录
        login_result = login(imageCode, cookies_josn, organ, username, session)

        # 登录成功 请求首页 获取_toke
        # if not login_result:
        _token=get_token_index(session)
        print("_token",_token)

        # 获取销售在途汇总表 2024年每月数据





        return login_result

    except Exception as e:
        print(f"[ERROR] 处理用户 {username} 时发生异常: {e}")
        return False
    finally:
        # 关闭Session
        session.close()
        print(f"[INFO] 用户 {username} 处理完成，Session已关闭")


def create_sales_report_excel(data_list, start_date="2024-1-1", end_date="2024-2-1",
                              filename="销售在途汇总表.xlsx"):
    """
    创建销售汇总Excel报表

    Args:
        data_list: 数据列表
        start_date: 开始日期
        end_date: 结束日期
        filename: 输出文件名
    """
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "销售在途汇总表"

    # 设置样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    subtitle_font = Font(name='微软雅黑', size=12, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='微软雅黑', size=10)

    # 居中对齐
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头填充色
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    # 第一行：主标题
    ws.merge_cells('A1:L1')
    ws['A1'] = f"{start_date}至{end_date} 销售在途汇总表"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    # 第二行：副标题
    current_date = datetime.now().strftime('%Y-%m-%d')
    ws.merge_cells('A2:L2')
    ws['A2'] = f"采集时间 {current_date} 销售在途汇总表"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = center_alignment

    # 第三行：表头
    headers = [
        '业务机构', '客户单位', '品种数', '数量', '码洋(元)',
        '含税金额/实洋(元)', '折扣额(元)', '税额(元)',
        '不含税金额(元)', '未开票数量', '未开票金额', '未开票成本'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # 填充数据
    for row_idx, data in enumerate(data_list, 4):
        for col_idx, (key, value) in enumerate(data.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # 设置对齐方式
            if key in ['业务机构', '客户单位']:
                cell.alignment = left_alignment
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # 特殊格式化：合计行
            if data.get('业务机构') == '合计':
                cell.font = Font(name='微软雅黑', size=10, bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # 设置列宽
    column_widths = [25, 20, 8, 10, 12, 15, 12, 12, 15, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 设置行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 20

    # 冻结窗格（冻结标题行）
    ws.freeze_panes = 'A4'

    # 保存文件
    wb.save(filename)
    print(f"Excel文件已生成: {filename}")

    return filename


# 主程序
if __name__ == "__main__":
    # usernames = ["13798252655", "13822786191"]
    usernames = ["13798252655", "13822786191","13631180662"]

    for username in usernames:
        result = process_user(username)
        print(f"\n[最终结果] 用户 {username} 处理{'成功' if result else '失败'}")
        time.sleep(2)  # 添加延迟，避免请求过快
